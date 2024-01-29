import os
import pandas as pd
import numpy as np
import re
from pathlib import Path
import xlwings as xw
import openpyxl
from openpyxl.styles import Font,Color,Alignment,Side, Border
import sys

# batch_folders = "Batch1"
batch_folders = input("Enter Folder Name: ")

path = os.getcwd()
file_path = path+"/input_folder/"+batch_folders+"/"

files = os.listdir(file_path)
files.sort()

files_csv = [f for f in files if f[-4:] == ".csv"]

appended_data = []
for file in files_csv:
	book = pd.read_csv(file_path+file, index_col=False, delimiter=',', error_bad_lines=False)

	df = pd.DataFrame(book)
	df.memory_usage(index=False, deep=True)
	df = df.dropna(how='all')
	df = df.replace(np.nan, "")

	maximum_marks = df.iloc[:, 3].replace("Absent", 0).astype(int)
	highest_marks = maximum_marks.max()

	# marks_optain_column = df.columns[2]
	# res = marks_optain_column[marks_optain_column.find('(')+1:marks_optain_column.find(')')].split(":")
	# out_of_marks = res[1].strip()

	out_of_marks = df.iloc[:, 5].astype(int)

	split_filename = file.split("_")
	# print(split_filename)
	# quit()

	date_string = split_filename[3].strip()
	date_value = date_string.replace(".csv", "" )

	df['date'] = date_value.strip()
	df['type'] = split_filename[1].strip()
	df['topic'] = split_filename[2].strip()
	df['marks obtained'] = maximum_marks
	df['out of'] = out_of_marks
	df['highest marks'] = highest_marks
	df['student name'] = df.iloc[:, 0].astype(str)

	cols = df.columns.tolist()
	cols = ['date', 'student name', 'type', 'topic', 'marks obtained', 'out of', 'highest marks']
	df = df[cols]

	appended_data.append(df)
	
appended_data = pd.concat(appended_data)
appended_data = appended_data.sort_values(['student name','date'])
appended_data.columns = appended_data.columns.str.upper()
appended_data = appended_data.apply(lambda x: x.astype(str).str.upper())

for i, x in appended_data.groupby('STUDENT NAME'):
	output_folder_path = path+"/output_folder/"+batch_folders
	if not os.path.exists(output_folder_path):
		os.makedirs(output_folder_path)
	i = i.replace(" ", "_")
	p = os.path.join(os.getcwd()+"/output_folder/"+batch_folders, batch_folders+"_{}_result.xlsx".format(i.lower()).strip())

	x.to_excel(p, index=False,startrow=7)
	
#for alignment in excel
def set_border(ws, side=None, blank=True):
	wb = sheet._parent
	side = side if side else Side(border_style='thin', color='000000')
	for cell in sheet._cells.values():
		cell.border = Border(top=side, bottom=side, left=side, right=side)
		cell.alignment = Alignment(horizontal='center', vertical='center')
		
excelFiles = os.listdir(output_folder_path+"/")
	 
# For each excel file
test_path = output_folder_path+"/"

for i in range(0, len(excelFiles)):
	wb = openpyxl.load_workbook(test_path+excelFiles[i])
	sheet = wb.active
	side = Side(border_style='thin', color='000000')
	set_border(sheet, side)
	
	a = sheet['B9']
	b = a.value

	font_style = Font(name="Calibri",size=15,color="000000",bold=True)
	cell = sheet["A6"]
	cell.font = font_style
	cell.value = "STUDENT NAME: "+b
	
	sheet.delete_cols(idx = 2)
	sheet.column_dimensions['A'].width = 12.80
	sheet.column_dimensions['B'].width = 12.00
	sheet.column_dimensions['C'].width = 33.71
	sheet.column_dimensions['D'].width = 18.71
	sheet.column_dimensions['E'].width = 10.43
	sheet.column_dimensions['F'].width = 18
	sheet.merge_cells('A1:F1')
	sheet.merge_cells('A2:F2')
	sheet.merge_cells('A4:F4')
	sheet.merge_cells('A6:F6')
	font_style = Font(name = "Calibri", size = 26, color = "002060", bold = True)
	sheet.row_dimensions[1].height = 40
	cell = sheet["A1"]
	cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
	cell.font = font_style
	cell.value = "MANOJ SIR AND MANISH SIR'S"

	font_style = Font(name = "Calibri", size = 22, color = "002060", bold = True)
	sheet.row_dimensions[2].height = 30
	cell = sheet["A2"]
	cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
	cell.font = font_style
	cell.value = "MATHEMATICS PRIVATE TUITIONS(MMMPT)"

	font_style = Font(name = "Calibri", size = 14, color = "000000", bold = True)
	cell = sheet["A4"]
	cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
	cell.font = font_style
	a = sheet['A9']
	b = a.value
	a = []
	for col in sheet['A']:
		a.append(col.value)
	
	c = a[-1]
	cell.value = ("REPORT CARD (FROM {} TO {})").format(b,c)


	wb.save(test_path+excelFiles[i])
	print(excelFiles[i] + ' completed.')

# When the code finishes, close the program.
sys.exit()



#combining all sheets into a single worksheet
"""SOURCE_DIR='output'
excel_files=list(Path(SOURCE_DIR).glob('*.xlsx'))
combined_wb=xw.Book()
for excel_file in excel_files:
	wb=xw.Book(excel_file)
	for sheet in wb.sheets:
		sheet.api.Copy(After=combined_wb.sheets[0].api)
	wb.close()
combined_wb.sheets[0].delete()
combined_wb.save(f'all_worksheets.xlsx')
if len(combined_wb.app.books)==1:
	combined_wb.app.quit()
else:
	combined_wb.close()"""











